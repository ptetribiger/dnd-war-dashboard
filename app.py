
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import pandas as pd
import random

# ----------------------------
# App configuration & header
# ----------------------------
st.set_page_config(page_title="DND War Dashboard (Web)", layout="wide")
st.title("DND War Dashboard — Interactive Control Center")

st.markdown(
    """
Upload **DND.xlsm** (or .xlsx) and use the tabs & sidebar tools to manage your campaign.

**Notes**
- Macros (VBA) are preserved on download, but not executed here.
- Use **Save / Download** to export after edits.
"""
)

# ----------------------------
# Sidebar: Quick Actions
# ----------------------------
with st.sidebar:
    st.header("Quick Actions")
    # Dice roller (logs to Events)
    st.subheader("Dice Roller")
    die = st.selectbox("Die", ["d4","d6","d8","d10","d12","d20"], index=5)
    mod = st.number_input("Modifier", value=0, step=1)
    roll_now = st.button("Roll")
    # Income boost simulator
    st.subheader("Income Boost")
    income_boost_pct = st.slider("Boost % (what-if)", 0, 100, 10, step=5)
    st.caption("This simulator adjusts charted income in the Dashboard (not saved).")
    # Add Day helper (just logs a note)
    st.subheader("Day Tracker")
    add_day_note = st.text_input("Note for 'Add Days' event", "")
    add_day_click = st.button("Log Day Note")

# ----------------------------
# File uploader
# ----------------------------
uploaded = st.file_uploader(
    "Upload your workbook", type=["xlsm", "xlsx"], accept_multiple_files=False
)

if not uploaded:
    st.info("Upload your Excel file to begin.")
    st.stop()

# ----------------------------
# Load workbook (preserve VBA)
# ----------------------------
try:
    buf_in = BytesIO(uploaded.read())
    wb = load_workbook(buf_in, read_only=False, keep_vba=True, data_only=True)
except Exception as e:
    st.error(f"Failed to load workbook: {e}")
    st.stop()

# Track whether user changed data (for Save tab messaging)
if "edited_flag" not in st.session_state:
    st.session_state.edited_flag = False

# ----------------------------
# Helpers
# ----------------------------
def get_ws(wb, name: str):
    """Return worksheet or None if missing."""
    try:
        return wb[name]
    except KeyError:
        return None

def find_header_row(ws, key="RegionName"):
    """Find header row index by matching `key` in column A (case-insensitive)."""
    if ws is None:
        return None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower() == key.lower():
            return r
    return None

def make_unique_columns(cols):
    """
    Ensure all column names are strings, not empty, and unique:
    - Replace None/blank with 'ColN'
    - Deduplicate with ' (1)', ' (2)', ...
    """
    norm = [str(c).strip() if (c is not None and str(c).strip() != "") else None for c in cols]
    filled = [c if c is not None else f"Col{i+1}" for i, c in enumerate(norm)]
    seen = {}
    unique = []
    for c in filled:
        if c not in seen:
            seen[c] = 0
            unique.append(c)
        else:
            seen[c] += 1
            unique.append(f"{c} ({seen[c]})")
    return unique

def sheet_to_dataframe_by_key(ws, header_key="RegionName"):
    """
    Build a DataFrame from a sheet that uses a header row identified by `header_key` in col A.
    """
    if ws is None:
        return None, "Worksheet not found."
    h = find_header_row(ws, header_key)
    if h is None:
        return None, f"Header row not found (looking for '{header_key}' in column A)."
    headers = [ws.cell(row=h, column=c).value for c in range(1, ws.max_column + 1)]
    columns = make_unique_columns(headers)

    rows = []
    for r in range(h + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in row):
            break
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    return df, None

def sheet_to_dataframe_first_row(ws):
    """
    Build a DataFrame assuming row 1 is the header (used for Monsters in the updated workbook).
    """
    if ws is None:
        return None, "Worksheet not found."
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    columns = make_unique_columns(headers)
    rows = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue
        rows.append(row)
    df = pd.DataFrame(rows, columns=columns)
    return df, None

def safe(v):
    return "" if v in (None, "#N/A") else v

# ----------------------------
# Tabs
# ----------------------------
tab_dash, tab_terr, tab_recon, tab_mon, tab_upg, tab_events, tab_save = st.tabs(
    ["Dashboard", "Territories", "Recon", "Monsters", "Upgrades", "Events", "Save / Download"]
)

# ----------------------------
# Dashboard (label:value from WarDashboard + interactive charts)
# ----------------------------
with tab_dash:
    ws = get_ws(wb, "WarDashboard")
    data_map = {}
    if ws is None:
        st.warning("Sheet 'WarDashboard' not found.")
    else:
        for r in range(1, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if key:
                data_map[str(key)] = val

        c1, c2, c3 = st.columns(3)
        c1.metric("Total Control %", safe(data_map.get("Total Control %", 0)))
        c2.metric("Total Income/day", safe(data_map.get("Total Income/day", 0)))
        c3.metric("Counterattack Risk", safe(data_map.get("Counterattack Risk", 0)))

        st.write("**Next Attack:**", safe(data_map.get("Next Attack", "")))
        st.write("**Attack Type:**", safe(data_map.get("Attack Type", "")))

    # Bonus KPIs + charts from Territories
    ws_t = get_ws(wb, "Territories")
    if ws_t:
        df_terr, err = sheet_to_dataframe_by_key(ws_t, header_key="RegionName")
        if not err and not df_terr.empty:
            # Compute simple counts by TerritoryState
            if "TerritoryState" in df_terr.columns:
                state_counts = df_terr["TerritoryState"].fillna("Unknown").value_counts()
                st.subheader("Territory States")
                st.bar_chart(state_counts)

            # What-if income boost chart
            if "RegionName" in df_terr.columns and "IncomeOutput" in df_terr.columns:
                chart_df = df_terr[["RegionName","IncomeOutput"]].copy()
                chart_df["IncomeOutput (what-if)"] = chart_df["IncomeOutput"].fillna(0) * (1 + income_boost_pct/100.0)
                st.subheader("Income Output (with What‑If Boost)")
                st.bar_chart(chart_df.set_index("RegionName")[["IncomeOutput (what-if)"]])

            # Risk vs Control scatter (if available)
            cols_needed = {"CounterattackRisk","CurrentControl"}
            if cols_needed.issubset(set(df_terr.columns)):
                scatter_df = df_terr[["RegionName","CurrentControl","CounterattackRisk"]].dropna()
                st.subheader("Risk vs. Control")
                st.scatter_chart(scatter_df.set_index("RegionName"))

# ----------------------------
# Territories (filter + toggle view/edit + conditional color)
# ----------------------------
with tab_terr:
    ws = get_ws(wb, "Territories")
    if ws is None:
        st.warning("Sheet 'Territories' not found.")
    else:
        df_terr, err = sheet_to_dataframe_by_key(ws, header_key="RegionName")
        if err:
            st.warning(err)
        else:
            # Filters
            st.subheader("Filters")
            states = sorted(df_terr.get("TerritoryState", pd.Series()).dropna().unique().tolist())
            state_filter = st.multiselect("State", states, default=states)
            name_query = st.text_input("Search RegionName")
            max_risk = int(df_terr.get("CounterattackRisk", pd.Series([0])).fillna(0).max())
            risk_cap = st.slider("Max CounterattackRisk", 0, max(0, max_risk), max(0, max_risk))

            view_df = df_terr.copy()
            if states:
                view_df = view_df[view_df["TerritoryState"].fillna("Unknown").isin(state_filter)]
            if name_query:
                view_df = view_df[view_df["RegionName"].astype(str).str.contains(name_query, case=False, na=False)]
            if "CounterattackRisk" in view_df.columns:
                view_df = view_df[view_df["CounterattackRisk"].fillna(0) <= risk_cap]

            # Toggle view/edit
            st.subheader("View / Edit")
            edit_mode = st.checkbox("Enable editing", value=False, help="Turn on to edit values and write back to workbook.")

            if not edit_mode:
                # Conditional color by risk (view mode)
                styled = view_df.style.apply(
                    lambda s: ["background-color: #ffe5e5" if (isinstance(x,(int,float)) and x>=5) else "" 
                               for x in s] if s.name=="CounterattackRisk" else ["" for _ in s],
                    axis=0
                )
                st.dataframe(styled, use_container_width=True)
            else:
                edited = st.data_editor(
                    view_df,
                    use_container_width=True,
                    num_rows="dynamic",
                    column_config={
                        "TerritoryState": st.column_config.SelectboxColumn(options=states or ["Lost","Controlled"], help="Set state"),
                        "CounterattackMonster": st.column_config.TextColumn(help="e.g., Obsidian Foot Soldier"),
                        "HowmanyMonsters": st.column_config.NumberColumn(min_value=0, step=1),
                        "IncomeOutput": st.column_config.NumberColumn(step=1),
                        "CounterattackRisk": st.column_config.NumberColumn(min_value=0, step=1),
                        "CurrentControl": st.column_config.NumberColumn(min_value=0, step=1),
                    },
                )
                if st.button("Apply changes to workbook"):
                    # Write back to underlying rows (match by RegionName)
                    # Re-find header (defensive)
                    h2 = find_header_row(ws, "RegionName")
                    start_row = h2 + 1 if h2 else 2
                    # Build index from sheet RegionName -> row number
                    name_to_row = {}
                    for r in range(start_row, ws.max_row + 1):
                        nm = ws.cell(row=r, column=1).value
                        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
                        if all(v is None for v in row_vals):
                            break
                        if nm:
                            name_to_row[str(nm)] = r
                    # Write each edited row
                    for _, row in edited.iterrows():
                        nm = str(row.get("RegionName"))
                        if nm in name_to_row:
                            r = name_to_row[nm]
                            for c_idx, col_name in enumerate(edited.columns, start=1):
                                ws.cell(row=r, column=c_idx).value = row[col_name]
                    st.session_state.edited_flag = True
                    st.success("Territories updated.")

# ----------------------------
# Recon (filters + scatter)
# ----------------------------
with tab_recon:
    ws = get_ws(wb, "Recon")
    if ws is None:
        st.warning("Sheet 'Recon' not found.")
    else:
        df_recon, err = sheet_to_dataframe_by_key(ws, header_key="RegionName")
        if err:
            st.warning(err)
        else:
            # Filters
            st.subheader("Filters")
            lvl_min, lvl_max = st.slider("ReconLevel range", 0, int(df_recon["ReconLevel"].fillna(0).max() or 5), (0, int(df_recon["ReconLevel"].fillna(0).max() or 5)))
            acc_opts = sorted(df_recon["Accuracy"].fillna("Unknown").unique().tolist())
            acc_sel = st.multiselect("Accuracy", acc_opts, default=acc_opts)
            q = st.text_input("Search RegionName")

            view_df = df_recon.copy()
            if "ReconLevel" in view_df.columns:
                view_df = view_df[
                    (view_df["ReconLevel"].fillna(0) >= lvl_min) &
                    (view_df["ReconLevel"].fillna(0) <= lvl_max)
                ]
            if "Accuracy" in view_df.columns:
                view_df = view_df[view_df["Accuracy"].fillna("Unknown").isin(acc_sel)]
            if q:
                view_df = view_df[view_df["RegionName"].astype(str).str.contains(q, case=False, na=False)]

            st.dataframe(view_df, use_container_width=True)

            # Noise vs ReconLevel scatter
            if {"NoiseRange","ReconLevel"}.issubset(set(view_df.columns)):
                s_df = view_df[["RegionName","ReconLevel","NoiseRange"]].dropna()
                st.subheader("Noise vs ReconLevel")
                st.scatter_chart(s_df.set_index("RegionName"))

# ----------------------------
# Monsters (filters + details)
# ----------------------------
with tab_mon:
    ws = get_ws(wb, "Monsters")
    if ws is None:
        st.warning("Sheet 'Monsters' not found.")
    else:
        df_mon, err = sheet_to_dataframe_first_row(ws)
        if err:
            st.warning(err)
        else:
            st.subheader("Filters")
            cr_min = int(pd.to_numeric(df_mon.get("CR", pd.Series([0])), errors="coerce").fillna(0).min())
            cr_max = int(pd.to_numeric(df_mon.get("CR", pd.Series([0])), errors="coerce").fillna(0).max())
            cr_rng = st.slider("CR range", cr_min, cr_max if cr_max>=cr_min else cr_min, (cr_min, cr_max if cr_max>=cr_min else cr_min))
            type_query = st.text_input("Type contains (e.g., 'Humanoid', 'Monstrosity')")
            name_query = st.text_input("Search Name")

            view = df_mon.copy()
            # Filter CR numerically if present
            if "CR" in view.columns:
                scr = pd.to_numeric(view["CR"], errors="coerce").fillna(0)
                view = view[(scr >= cr_rng[0]) & (scr <= cr_rng[1])]
            # Type filter
            if "Size/Type" in view.columns and type_query:
                view = view[view["Size/Type"].astype(str).str.contains(type_query, case=False, na=False)]
            # Name filter
            if name_query:
                view = view[view["Name"].astype(str).str.contains(name_query, case=False, na=False)]

            st.dataframe(view, use_container_width=True)

            # Details panel
            st.subheader("Monster Details")
            pick = st.selectbox("Select a monster", options=view["Name"].astype(str).unique().tolist())
            if pick:
                mon_row = view[view["Name"].astype(str) == pick].iloc[0].to_dict()
                c1, c2 = st.columns(2)
                with c1:
                    st.write("**Core Stats**")
                    for k in ["CR","Size/Type","HP","Speed","STR","DEX","CON","INT","WIS","CHA"]:
                        if k in mon_row:
                            st.write(f"{k}: {mon_row[k]}")
                with c2:
                    st.write("**Traits & Actions**")
                    for k in ["Senses","Traits","Actions","Legendary Actions","Tactics"]:
                        if k in mon_row:
                            st.write(f"{k}: {mon_row[k]}")

# ----------------------------
# Upgrades: list tiers & event codes (unchanged, but grouped)
# ----------------------------
with tab_upg:
    ws = get_ws(wb, "Upgrade Systems")
    if ws is None:
        st.warning("Sheet 'Upgrade Systems' not found.")
    else:
        weapons, militia, event_types = [], [], []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            if not isinstance(a, str):
                continue
            text = a.strip()
            if text.lower().startswith("weapon tier"):
                block = [ws.cell(row=rr, column=1).value for rr in range(r, r + 8)]
                weapons.append("; ".join(str(x) for x in block if x))
            elif text.lower().startswith("militia tier"):
                block = [ws.cell(row=rr, column=1).value for rr in range(r, r + 8)]
                militia.append("; ".join(str(x) for x in block if x))
            elif text.isupper() and "_" in text:
                event_types.append(text)

        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Weapon tiers")
            if weapons:
                for w in weapons:
                    st.write("- ", w)
            else:
                st.info("(none detected)")

        with col2:
            st.subheader("Militia tiers")
            if militia:
                for m in militia:
                    st.write("- ", m)
            else:
                st.info("(none detected)")

        st.subheader("Event codes detected")
        st.write(", ".join(sorted(set(event_types))) or "(none)")

# ----------------------------
# Events (form + sidebar logging integration)
# ----------------------------
with tab_events:
    # Regions for dropdown from Territories
    ws_t = get_ws(wb, "Territories")
    regions = []
    if ws_t:
        h = find_header_row(ws_t, "RegionName")
        if h:
            for r in range(h + 1, ws_t.max_row + 1):
                name = ws_t.cell(row=r, column=1).value
                rest = [ws_t.cell(row=r, column=c).value for c in range(1, ws_t.max_column + 1)]
                if all(v is None for v in rest):
                    break
                if name:
                    regions.append(str(name))

    # Event type list from Upgrade Systems
    ws_u = get_ws(wb, "Upgrade Systems")
    event_types = []
    if ws_u:
        for r in range(1, ws_u.max_row + 1):
            a = ws_u.cell(row=r, column=1).value
            if isinstance(a, str):
                t = a.strip()
                if t.isupper() and "_" in t:
                    event_types.append(t)

    # Ensure TerritoryEvents exists with header
    ws_e = get_ws(wb, "TerritoryEvents")
    if not ws_e:
        ws_e = wb.create_sheet("TerritoryEvents")
        ws_e.cell(row=1, column=1).value = "RegionName"
        ws_e.cell(row=1, column=2).value = "EventType"
        ws_e.cell(row=1, column=3).value = "Value"
        ws_e.cell(row=1, column=4).value = "Notes"

    st.subheader("Add an Event")
    with st.form("add_event_form", clear_on_submit=True):
        region = st.selectbox("Region", options=regions or [""], index=0 if regions else None)
        etype = st.selectbox("Event Type", options=sorted(set(event_types)) or [""], index=0 if event_types else None)
        value = st.number_input("Value", value=0.0, step=1.0)
        notes = st.text_input("Notes", value="")
        submitted = st.form_submit_button("Add Event")
        if submitted:
            target_row = ws_e.max_row + 1
            ws_e.cell(row=target_row, column=1).value = region
            ws_e.cell(row=target_row, column=2).value = etype
            ws_e.cell(row=target_row, column=3).value = float(value)
            ws_e.cell(row=target_row, column=4).value = notes
            st.session_state.edited_flag = True
            st.success(f"Event appended: {etype} in {region} (Value {value}).")

    # Sidebar actions: dice roller & day note -> append events automatically
    if roll_now:
        # Parse die like d20
        sides = int(die[1:])
        result = random.randint(1, sides) + int(mod)
        target_row = ws_e.max_row + 1
        ws_e.cell(row=target_row, column=1).value = safe(st.session_state.get("last_region") or "GLOBAL")
        ws_e.cell(row=target_row, column=2).value = f"ROLL_{die.upper()}"
        ws_e.cell(row=target_row, column=3).value = result
        ws_e.cell(row=target_row, column=4).value = f"Roll {die} + {mod} = {result}"
        st.session_state.edited_flag = True
        st.toast(f"🎲 Rolled {die} + {mod} → **{result}**", icon="🎲")

    if add_day_click and add_day_note:
        target_row = ws_e.max_row + 1
        ws_e.cell(row=target_row, column=1).value = safe(st.session_state.get("last_region") or "GLOBAL")
        ws_e.cell(row=target_row, column=2).value = "ADD_DAYS_NOTE"
        ws_e.cell(row=target_row, column=3).value = 0.0
        ws_e.cell(row=target_row, column=4).value = add_day_note
        st.session_state.edited_flag = True
        st.toast("🗓️ Day note logged.", icon="🗒️")

# ----------------------------
# Save & Download (preserve macros)
# ----------------------------
with tab_save:
    changed = st.session_state.edited_flag
    if changed:
        st.success("You have unsaved changes. Use the button below to download.")
    else:
        st.info("No changes detected. You can still download a copy if needed.")

    st.write("Download the updated workbook (macros preserved if present).")
    try:
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        # choose MIME based on extension
        is_xlsm = uploaded.name.lower().endswith(".xlsm")
        mime = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if is_xlsm
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            "Download updated workbook",
            data=out,
            file_name=uploaded.name,
            mime=mime,
        )
    except Exception as e:
        st.error(f"Failed to prepare download: {e}")
